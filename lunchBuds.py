import pandas as pd
import openpyxl
import random
from queue import Queue

path = "LunchBuds.xlsx"

df = pd.read_excel(path)
df.head()
allMems = df["Names"]
nameMap = {}
justNames = []
eboards = set([""]) #put eboard names here
totalWeeks = len(df.columns)

for name in allMems:
    justNames.append(name.replace(u'\xa0', u''))

for row in range(len(df)):
    currSet = set(df.iloc[row][1:totalWeeks].values.tolist())
    currPerson = justNames[row]
    nameMap[currPerson] = currSet

result = {}
matched = set()

for idx in range(len(justNames)):
    available = []
    currMem = justNames[idx]

    if currMem in matched:
        continue

    if currMem in eboards:
        available = [x for x in justNames if x not in eboards and x not in nameMap[currMem] and x != currMem and x not in matched]
    else:
        available = [x for x in justNames if x not in nameMap[currMem] and x != currMem and x not in matched]
    random.shuffle(available)
    for match in available:
        if match in matched:
            continue
        else:
            matched.add(currMem)
            matched.add(match)
            result[currMem] = match
            break

column = []
newResult = {}

for r in result:
    value = result[r] 
    newResult[r] = value
    newResult[value] = r

for mem in justNames:
    column.append(newResult[mem])    

currRow = 1

try:
    wb = openpyxl.load_workbook("LunchBuds.xlsx")
    ws = wb.worksheets[0]
    colNum = totalWeeks + 1
    title = "Week " + str(totalWeeks)
    ws.cell(row=currRow, column=colNum).value = title
    currRow += 1
    for c in column:
        ws.cell(row=currRow, column=colNum).value = c
        currRow += 1
    wb.save("LunchBuds.xlsx")
except:
    raise SystemExit(0)

for r in result:
    print(r + " <-> " + result[r])
    # print empty line
    print()